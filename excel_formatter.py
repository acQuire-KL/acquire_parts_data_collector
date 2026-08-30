"""Excel-only presentation helpers for PDC workbook outputs."""

from __future__ import annotations

from datetime import datetime
import re

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from excel_formats import format_for_heading

PROVIDER_BLOCK_COLOURS = [
    "D9EAF7",  # pale blue
    "E2F0D9",  # pale green
    "FFF2CC",  # pale yellow
    "FCE4D6",  # pale peach
    "E4DFEC",  # pale lavender
    "E7E6E6",  # pale grey
    "DDEBF7",  # pale aqua
    "E2EFDA",  # pale mint
    "F4CCCC",  # pale rose
    "F3E5AB",  # pale beige
]

GROUP_COLOURS = {
    "Input & Match": "5B9BD5",
    "Identity": "4472C4",
    "Documentation": "70AD47",
    "Compliance": "8064A2",
    "Physical": "A5A5A5",
    "Electrical": "ED7D31",
    "Commercial (Existing)": "FFC000",
    "Commercial": "FFC000",
    "Traceability": "264478",
}

STATUS_FILLS = {
    "Matched": "C6EFCE",             # green
    "Review Required": "FFEB9C",     # yellow
    "Multiple Matches": "F4B183",    # orange
    "Not Found": "FFC7CE",           # red
}

STATUS_FONTS = {
    "Matched": "006100",
    "Review Required": "9C6500",
    "Multiple Matches": "9C5700",
    "Not Found": "9C0006",
}

ENRICHED_PARTS_FREEZE_PANES = "E3"

_NUMERIC_TEXT = re.compile(r"^[+-]?(?:\d+(?:,\d{3})*|\d*)(?:\.\d+)?$")


def group_colour(group: str) -> str:
    """Return a group colour, cycling the provider-position palette as needed."""
    if group.startswith("Provider #"):
        try:
            position = int(group.split("#", 1)[1].strip())
        except (ValueError, IndexError):
            position = 1
        return PROVIDER_BLOCK_COLOURS[(position - 1) % len(PROVIDER_BLOCK_COLOURS)]
    return GROUP_COLOURS.get(group, "1F4E78")


def _column_group(column) -> str:
    if hasattr(column, "group"):
        return str(column.group)
    return str(column[0])


def add_group_headers(ws, columns) -> None:
    """Add the merged top-level group row used by review-oriented sheets."""
    column_index = 1
    while column_index <= len(columns):
        group = _column_group(columns[column_index - 1])
        start = column_index
        while column_index <= len(columns) and _column_group(columns[column_index - 1]) == group:
            column_index += 1
        end = column_index - 1
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        cell = ws.cell(1, start, group)
        fill_colour = group_colour(group)
        cell.fill = PatternFill("solid", fgColor=fill_colour)
        font_colour = "000000" if group.startswith("Provider #") else "FFFFFF"
        cell.font = Font(color=font_colour, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _coerce_numeric(value):
    """Convert clean numeric strings to numbers without touching identifiers."""
    if value in (None, "") or isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value

    text = str(value).strip()
    if not text or not _NUMERIC_TEXT.fullmatch(text):
        return value

    number = float(text.replace(",", ""))
    return int(number) if number.is_integer() else number


def _coerce_date(value):
    """Convert ISO-style date text when possible; otherwise preserve the source."""
    if value in (None, "") or isinstance(value, datetime):
        return value
    text = str(value).strip()
    if not text:
        return value
    try:
        return datetime.fromisoformat(text.replace("Z", "+00:00")).replace(tzinfo=None)
    except ValueError:
        return value


def _set_column_widths(ws, header_row: int) -> None:
    for column in range(1, ws.max_column + 1):
        heading = str(ws.cell(header_row, column).value or "")
        field_format = format_for_heading(heading)

        if field_format.width is not None:
            width = field_format.width
        else:
            max_length = 0
            for row in range(header_row, ws.max_row + 1):
                max_length = max(max_length, len(str(ws.cell(row, column).value or "")))
            width = min(max(max_length + 2, 10), 40)

        ws.column_dimensions[get_column_letter(column)].width = width


def _apply_field_formats(ws, headings: list[str], first_data_row: int) -> None:
    for column, heading in enumerate(headings, 1):
        field_format = format_for_heading(heading)

        for row_number in range(first_data_row, ws.max_row + 1):
            cell = ws.cell(row_number, column)

            if field_format.coerce_numeric:
                cell.value = _coerce_numeric(cell.value)
            elif field_format.number_format in {"yyyy-mm-dd", "yyyy-mm-dd hh:mm:ss"}:
                cell.value = _coerce_date(cell.value)

            cell.number_format = field_format.number_format
            if field_format.font_name:
                cell.font = Font(name=field_format.font_name)
            cell.alignment = Alignment(
                horizontal=field_format.horizontal,
                vertical=field_format.vertical,
                wrap_text=field_format.wrap_text,
            )

            if field_format.hyperlink and str(cell.value or "").startswith("http"):
                cell.hyperlink = str(cell.value)
                cell.style = "Hyperlink"
                # Reapply alignment because the Hyperlink named style can alter it.
                cell.alignment = Alignment(
                    horizontal=field_format.horizontal,
                    vertical=field_format.vertical,
                    wrap_text=field_format.wrap_text,
                )



def _estimated_wrapped_lines(value, column_width: float | None) -> int:
    """Estimate Excel display lines from explicit breaks and column width.

    openpyxl cannot ask Excel to AutoFit a row.  This approximation counts
    both embedded newlines and text that will wrap visually within the
    configured column width.
    """
    if value in (None, ""):
        return 1

    # Excel column width is approximately a character count for the default
    # font.  Leave a little margin so the estimate does not clip at the edge.
    usable_characters = max(1, int((column_width or 10) * 0.90))
    visual_lines = 0
    for explicit_line in str(value).split("\n"):
        visual_lines += max(1, (len(explicit_line) + usable_characters - 1) // usable_characters)
    return visual_lines


def _set_wrapped_row_heights(ws, headings: list[str], first_data_row: int) -> None:
    """Size rows for the tallest wrapped cell, including full price ladders."""
    wrapped_columns = [
        column
        for column, heading in enumerate(headings, 1)
        if format_for_heading(heading).wrap_text
    ]
    if not wrapped_columns:
        return

    for row_number in range(first_data_row, ws.max_row + 1):
        line_count = 1
        for column in wrapped_columns:
            letter = get_column_letter(column)
            width = ws.column_dimensions[letter].width
            line_count = max(
                line_count,
                _estimated_wrapped_lines(ws.cell(row_number, column).value, width),
            )

        # 15 points is a practical single-line height for the workbook font.
        # The higher ceiling allows complete multi-offer price ladders while
        # still protecting the sheet from pathological free-text values.
        ws.row_dimensions[row_number].height = min(max(18, line_count * 15), 420)


def _apply_status_colours(ws, headings: list[str], first_data_row: int) -> None:
    if "Match Status" not in headings:
        return
    status_column = headings.index("Match Status") + 1
    for row_number in range(first_data_row, ws.max_row + 1):
        cell = ws.cell(row_number, status_column)
        status = str(cell.value or "")
        fill_colour = STATUS_FILLS.get(status)
        if fill_colour:
            cell.fill = PatternFill("solid", fgColor=fill_colour)
            cell.font = Font(color=STATUS_FONTS[status], bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="top")


def _wrap_all_cells(ws) -> None:
    """4.7.2b: make all workbook output cells readable without manual wrapping."""
    for row in ws.iter_rows():
        for cell in row:
            current = cell.alignment
            horizontal = current.horizontal
            if str(cell.value or "") and cell.column <= ws.max_column:
                # Lead Time (Weeks) is deliberately numeric and centred.
                heading_row = 2 if ws.title in ("Enriched Parts", "Review Required") else 1
                heading = str(ws.cell(heading_row, cell.column).value or "")
                if heading == "Lead Time (Weeks)":
                    horizontal = "center"
            cell.alignment = Alignment(
                horizontal=horizontal,
                vertical=current.vertical or "top",
                text_rotation=current.text_rotation,
                wrap_text=True,
                shrink_to_fit=current.shrink_to_fit,
                indent=current.indent,
            )


def format_review_sheet(ws, headings: list[str]) -> None:
    """Format Enriched Parts or Review Required for interactive review."""
    ws.freeze_panes = ENRICHED_PARTS_FREEZE_PANES
    ws.auto_filter.ref = f"A2:{get_column_letter(ws.max_column)}{ws.max_row}"

    for cell in ws[2]:
        group = ws.cell(1, cell.column).value
        cell.font = Font(color="FFFFFF", bold=True)
        fill_colour = group_colour(str(group or ""))
        cell.fill = PatternFill("solid", fgColor=fill_colour)
        if str(group or "").startswith("Provider #"):
            cell.font = Font(color="000000", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 36

    _apply_field_formats(ws, headings, first_data_row=3)
    _set_column_widths(ws, header_row=2)
    _set_wrapped_row_heights(ws, headings, first_data_row=3)
    _apply_status_colours(ws, headings, first_data_row=3)
    _wrap_all_cells(ws)


def format_reference_sheet(ws) -> None:
    """Apply basic presentation without filters or frozen panes."""
    headings = [str(cell.value or "") for cell in ws[1]]

    for cell in ws[1]:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    _apply_field_formats(ws, headings, first_data_row=2)
    _set_column_widths(ws, header_row=1)
    _set_wrapped_row_heights(ws, headings, first_data_row=2)
    _wrap_all_cells(ws)
