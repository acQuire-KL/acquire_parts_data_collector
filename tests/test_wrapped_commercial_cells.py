import unittest

from openpyxl import Workbook

from excel_formatter import add_group_headers, format_review_sheet
from excel_formats import format_for_heading


class WrappedCommercialCellsTests(unittest.TestCase):
    def test_multi_offer_headings_are_wrapped(self):
        for heading in (
            "Provider Part Number",
            "Pack Format",
            "Packaging Code",
            "Minimum Order Quantity",
            "Pack Quantity",
            "Quantity Available",
            "Additional Charge",
            "Additional Charge Description",
            "Price Breaks",
        ):
            self.assertTrue(format_for_heading(heading).wrap_text, heading)

    def test_review_sheet_displays_three_line_offer_values(self):
        wb = Workbook()
        ws = wb.active
        columns = [
            ("Commercial", "Provider Part Number"),
            ("Commercial", "Pack Format"),
        ]
        add_group_headers(ws, columns)
        headings = [heading for _, heading in columns]
        ws.append(headings)
        ws.append([
            "Cut Tape: 535-13402-1-ND\nDigiReel: 535-13402-6-ND\nReel: 535-13402-2-ND",
            "Cut Tape: Cut Tape\nDigiReel: DigiReel\nReel: Reel",
        ])

        format_review_sheet(ws, headings)

        self.assertTrue(ws["A3"].alignment.wrap_text)
        self.assertEqual(ws["A3"].alignment.vertical, "top")
        self.assertGreaterEqual(ws.row_dimensions[3].height, 45)


if __name__ == "__main__":
    unittest.main()
