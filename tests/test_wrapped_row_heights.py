import unittest

from openpyxl import Workbook

from excel_formatter import add_group_headers, format_review_sheet


class WrappedRowHeightTests(unittest.TestCase):
    def test_full_price_ladder_is_not_capped_at_six_lines(self):
        wb = Workbook()
        ws = wb.active
        columns = [("Commercial", "Price Breaks")]
        add_group_headers(ws, columns)
        headings = ["Price Breaks"]
        ws.append(headings)
        price_ladder = "\n".join(f"{quantity:,} @ 0.12345" for quantity in range(1, 28))
        ws.append([price_ladder])

        format_review_sheet(ws, headings)

        self.assertTrue(ws["A3"].alignment.wrap_text)
        self.assertGreaterEqual(ws.row_dimensions[3].height, 405)

    def test_long_unbroken_text_also_increases_height(self):
        wb = Workbook()
        ws = wb.active
        columns = [("Commercial", "Provider Part Number")]
        add_group_headers(ws, columns)
        headings = ["Provider Part Number"]
        ws.append(headings)
        ws.append(["X" * 100])

        format_review_sheet(ws, headings)

        self.assertGreater(ws.row_dimensions[3].height, 18)


if __name__ == "__main__":
    unittest.main()
