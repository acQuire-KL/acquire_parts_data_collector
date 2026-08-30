import argparse
import csv
from collections import OrderedDict
from datetime import datetime, timezone
from concurrent.futures import ThreadPoolExecutor
from time import perf_counter
from pathlib import Path

from openpyxl import Workbook, load_workbook

from config import Settings, MouserSettings, TmeSettings
from providers import ProviderManager
from providers.digikey import DigiKeyProvider
from providers.mouser import MouserProvider
from providers.tme import TmeProvider
from manufacturer_resolver import names_equivalent, resolve_manufacturer
from excel_formatter import add_group_headers, format_reference_sheet, format_review_sheet
from workbook_layout import WorkbookColumn, column_keys, display_headings, enriched_parts_columns
from commercial_profile import commercial_offers
from knowledge_base_manager import KnowledgeBaseManager
from operational_bom_review import PartsMasterLookup, provider_review_observation, summary_rows
from attribute_normalization import normalise_attribute
from lead_time_normalization import lead_time_display, normalise_lead_time
from identity_recovery import (
    RecoveryCandidate, candidate_from_profile, consolidate_candidates,
    discover_payload_candidates, footprint_consistency, normalise_mpn, search_variants, family_search_variants,
    recover_mpn_from_bom,
)
from multi_provider_summary import (
    ProviderEvidence, engineering_confirmation, evidence_status, merged_value,
    provider_availability, provider_identity_match,
)

APP_VERSION = "0.2.10 / Sprint 4.7.2c"

MFG = {"manufacturer", "mfg", "mf", "mfr", "manufacturer name"}
MPN = {"mpn", "manufacturer part number", "mfg part number", "manufacturer_part_number"}
DESCRIPTION = {"description", "desc", "part description"}
VALUE = {"value", "part value"}
FOOTPRINT = {"footprint", "package", "package / case"}
QUANTITY = {"quantity", "qty", "qty per", "bom qty"}
DNP = {"dnp", "do not populate", "do not fit", "fit", "fitted"}


def clean(value):
    return " ".join(str(value or "").strip().lower().replace("_", " ").split())


def norm(value):
    return "".join(c for c in str(value or "").upper() if c.isalnum())


def ci(data, *names):
    if not isinstance(data, dict):
        return ""
    lookup = {str(k).lower(): v for k, v in data.items()}
    for item in names:
        if item.lower() in lookup:
            return lookup[item.lower()]
    return ""


def name(value):
    if isinstance(value, dict):
        return str(ci(value, "Name", "Value", "ProductDescription", "Description") or "")
    return str(value or "")


def product(payload):
    return payload.get("Product") or payload.get("product") or payload


def descriptions(p):
    description = ci(p, "Description")
    if isinstance(description, dict):
        return (
            str(ci(description, "ProductDescription", "Description") or ""),
            str(ci(description, "DetailedDescription", "DetailedProductDescription") or ""),
        )
    return name(description), name(ci(p, "DetailedDescription", "DetailedProductDescription"))


def params(p):
    result = {}
    for item in (p.get("Parameters") or p.get("parameters") or []):
        if isinstance(item, dict):
            key = name(ci(item, "ParameterText", "Parameter", "Name")).strip().lower()
            value = name(ci(item, "ValueText", "Value"))
            if key:
                result[key] = value
    return result


def parameter_value(parameters, *aliases):
    for alias in aliases:
        value = parameters.get(alias.lower(), "")
        if value not in ("", "-"):
            return value
    return ""


def category_names(category):
    """Return top-level category and the deepest available child category."""
    if not isinstance(category, dict):
        return "", ""
    top = str(category.get("Name") or "")
    deepest = top
    current = category
    while isinstance(current, dict):
        children = current.get("ChildCategories") or []
        if not children or not isinstance(children[0], dict):
            break
        current = children[0]
        deepest = str(current.get("Name") or deepest)
    return top, deepest


def normalise_url(value):
    text = str(value or "").strip()
    if text.startswith("//"):
        return "https:" + text
    return text


def flatten(value, prefix=""):
    if isinstance(value, dict):
        for key, item in value.items():
            yield from flatten(item, f"{prefix}.{key}" if prefix else str(key))
    elif isinstance(value, list):
        for index, item in enumerate(value):
            yield from flatten(item, f"{prefix}[{index}]")
    else:
        yield prefix, value


def load_input_workbook(path, *, data_only=False):
    """Load XLSX/XLSM or CSV input into an openpyxl-compatible workbook.

    CSV is converted in memory so the operational review path can use the
    same downstream column and row handling as Excel input.
    """
    source_path = Path(path)
    if source_path.suffix.casefold() != ".csv":
        return load_workbook(source_path, data_only=data_only)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "CSV Input"
    with source_path.open("r", encoding="utf-8-sig", newline="") as handle:
        sample = handle.read(4096)
        handle.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        for row in csv.reader(handle, dialect):
            sheet.append(row)
    return workbook


def _normalised_provider_status(status, message=""):
    """Convert provider execution detail to concise operational states.

    Raw provider messages remain attached to ProviderEvidence for Knowledge
    Base diagnostics; only the status is normalised for review logic.
    """
    value = str(status or "").strip().casefold()
    text = str(message or "").casefold()
    if value == "error" and any(token in text for token in (
        "404", "not found", "no product", "no match", "no results",
        "zero results", "not listed", "does not list",
    )):
        return "no_match"
    return value or "error"


def _concise_provider_status(item):
    status = _normalised_provider_status(item.execution_status, item.message)
    text = str(item.message or "").casefold()
    if status == "success":
        return "success"
    if status == "no_match":
        return "not listed"
    if status == "skipped":
        if "mpn" in text and ("missing" in text or "blank" in text):
            return "skipped - MPN missing"
        return "skipped"
    if any(token in text for token in ("401", "403", "unauthorized", "forbidden", "authentication")):
        return "authentication error"
    if any(token in text for token in ("429", "rate limit", "too many requests")):
        return "rate limited"
    return "provider error"


def _progress(text):
    print(text, flush=True)


def locate_columns(headers):
    manufacturer_column = mpn_column = None
    for index, header in enumerate(headers, 1):
        value = clean(header)
        if value in MFG:
            manufacturer_column = index
        if value in MPN:
            mpn_column = index
    if not manufacturer_column or not mpn_column:
        raise ValueError(f"Could not identify Manufacturer and MPN columns: {headers}")
    return manufacturer_column, mpn_column


def locate_optional_column(headers, aliases):
    for index, header in enumerate(headers, 1):
        if clean(header) in aliases:
            return index
    return None


def input_cell(sheet, row, column):
    if not column:
        return ""
    value = sheet.cell(row, column).value
    return "" if value is None else value


def enriched_columns():
    """Return the configured Enriched Parts workbook layout."""
    return enriched_parts_columns()


def first_variation_value(p, *names):
    variations = p.get("ProductVariations") or []
    for variation in variations:
        if not isinstance(variation, dict):
            continue
        value = ci(variation, *names)
        if value not in ("", None):
            return value
    return ""




def _numeric_sort(value):
    try:
        return (0, float(value))
    except (TypeError, ValueError):
        return (1, str(value or ""))


def primary_commercial_offer(profile):
    """Choose a review-friendly offer while retaining every offer in Commercial Analysis."""
    variations = commercial_offers(profile)
    if not variations:
        return {}
    for preferred in ("Cut Tape", "Loose", "Tube", "Tray", "Reel", "DigiReel"):
        for variation in variations:
            if variation.get("pack_format") == preferred:
                return variation
    return variations[0]


def _price_break_texts(breaks):
    """Return sorted, formatted quantity/price pairs for one price ladder."""
    items = list(breaks or [])
    items.sort(key=lambda item: _numeric_sort(item.get("break_quantity")))
    formatted = []
    for item in items:
        quantity = item.get("break_quantity")
        price = item.get("unit_price")
        try:
            quantity_text = f"{int(float(quantity)):,}"
        except (TypeError, ValueError):
            quantity_text = str(quantity or "")
        try:
            price_text = f"{float(price):,.5f}"
        except (TypeError, ValueError):
            price_text = str(price or "")
        formatted.append((quantity_text, price_text))
    return formatted


def price_break_summary(variation, quantity_width=None, price_width=None):
    """Return an aligned multiline price ladder without a currency symbol."""
    formatted = _price_break_texts((variation or {}).get("standard_price_breaks") or [])
    if not formatted:
        return ""
    quantity_width = quantity_width or max(len(quantity) for quantity, _ in formatted)
    price_width = price_width or max(len(price) for _, price in formatted)
    return "\n".join(
        f"{quantity:>{quantity_width}}  {price:>{price_width}}"
        for quantity, price in formatted
    )


def first_additional_charge(variation):
    charges = list((variation or {}).get("additional_charges") or [])
    return charges[0] if charges else {}


def commercial_analysis_rows(result, profile):
    rows = []
    for variation in commercial_offers(profile):
        charge = first_additional_charge(variation)
        price_breaks = variation.get("standard_price_breaks") or []
        if not price_breaks:
            price_breaks = [{}]
        for price_break in price_breaks:
            rows.append(OrderedDict([
                ("Source Row", result.get("Source Row", "")),
                ("Manufacturer", result.get("Manufacturer", "")),
                ("Manufacturer Part Number", result.get("Manufacturer Part Number", "")),
                ("Provider", profile.get("provider", "")),
                ("Provider Part Number", variation.get("provider_part_number", "")),
                ("Currency", profile.get("provider_currency", "")),
                ("Pack Format", variation.get("pack_format", "")),
                ("Packaging Code", variation.get("packaging_code", variation.get("package_type", ""))),
                ("Minimum Order Quantity", variation.get("minimum_order_quantity", "")),
                ("Pack Quantity", variation.get("pack_quantity", "")),
                ("Quantity Available", variation.get("quantity_available", "")),
                ("Manufacturer Lead Weeks", lead_time_display(profile.get("manufacturer_lead_weeks", ""))),
                ("Break Quantity", price_break.get("break_quantity", "")),
                ("Unit Price", price_break.get("unit_price", "")),
                ("Extended Price", price_break.get("total_price", "")),
                ("Additional Charge", charge.get("amount", "")),
                ("Additional Charge Currency", charge.get("currency", profile.get("provider_currency", ""))),
                ("Additional Charge Description", charge.get("description", "")),
                ("Additional Charge Application", charge.get("application", "")),
                ("Captured At UTC", profile.get("captured_at_utc", "")),
            ]))
    return rows


def offer_value_summary(profile, field, fallback_field=""):
    """Return one labelled line per commercial offer for a selected field."""
    lines = []
    for offer in commercial_offers(profile):
        label = offer.get("pack_format") or offer.get("package_type") or "Offer"
        value = offer.get(field)
        if value in (None, "") and fallback_field:
            value = offer.get(fallback_field)
        if value not in (None, ""):
            lines.append(f"{label}: {value}")
    return "\n".join(lines)


def all_additional_charges_summary(profile, field):
    lines = []
    for offer in commercial_offers(profile):
        label = offer.get("pack_format") or offer.get("package_type") or "Offer"
        for charge in offer.get("additional_charges") or []:
            value = charge.get(field)
            if value not in (None, ""):
                lines.append(f"{label}: {value}")
    return "\n".join(lines)


def all_price_breaks_summary(profile):
    """Return all offer ladders aligned to common quantity and price widths."""
    offers = []
    all_pairs = []
    for offer in commercial_offers(profile):
        pairs = _price_break_texts(offer.get("standard_price_breaks") or [])
        if pairs:
            offers.append((offer, pairs))
            all_pairs.extend(pairs)
    if not all_pairs:
        return ""

    quantity_width = max(len(quantity) for quantity, _ in all_pairs)
    price_width = max(len(price) for _, price in all_pairs)
    sections = []
    for offer, _ in offers:
        label = offer.get("pack_format") or offer.get("package_type") or "Offer"
        summary = price_break_summary(offer, quantity_width, price_width)
        sections.append(f"{label}\n{summary}")
    return "\n\n".join(sections)


def build_result(row, requested_mfg, requested_mpn, p, pa, record, resolved):
    returned_mfg = name(ci(p, "Manufacturer"))
    returned_mpn = str(ci(p, "ManufacturerProductNumber", "ManufacturerPartNumber", "MfrPartNumber") or "")
    mpn_match = norm(requested_mpn) == norm(returned_mpn)
    returned_mfg_id = ci(ci(p, "Manufacturer"), "Id")
    id_match = bool(returned_mfg_id) and str(returned_mfg_id) == str(resolved.manufacturer_id)
    name_match = names_equivalent(resolved.matched_name, returned_mfg) or names_equivalent(requested_mfg, returned_mfg)
    manufacturer_match = id_match or name_match
    status = "Matched" if mpn_match and manufacturer_match else "Review Required"

    if status == "Matched":
        verification = "manufacturer ID" if id_match else "normalised manufacturer name"
        reason = (
            f"Exact normalised MPN; manufacturer verified by {verification}. "
            f"Input {requested_mfg!r} resolved to {returned_mfg!r}."
        )
    elif not mpn_match:
        reason = f"Returned MPN {returned_mpn!r} differs from requested MPN {requested_mpn!r}."
    else:
        reason = (
            f"Returned manufacturer {returned_mfg!r} does not match resolved manufacturer "
            f"{resolved.matched_name!r} (DigiKey ID {resolved.manufacturer_id})."
        )

    short_description, detailed_description = descriptions(p)
    category, family = category_names(ci(p, "Category"))
    classifications = ci(p, "Classifications")
    product_status = ci(p, "ProductStatus", "Status")
    commercial_profile = record.commercial_profile or {}
    primary_offer = primary_commercial_offer(commercial_profile)
    additional_charge = first_additional_charge(primary_offer)

    values = OrderedDict([
        ("Source Row", row),
        ("Requested Manufacturer", requested_mfg),
        ("Requested MPN", requested_mpn),
        ("Match Status", status),
        ("Reason", reason),
        ("Manufacturer", returned_mfg),
        ("Manufacturer Part Number", returned_mpn),
        ("DigiKey Part Number", first_variation_value(p, "DigiKeyProductNumber", "DigiKeyPartNumber", "ProductNumber")),
        ("Description", short_description),
        ("Detailed Description", detailed_description),
        ("Product Category", category),
        ("Product Family", family),
        ("Series", name(ci(p, "Series"))),
        ("Base Product Number", name(ci(p, "BaseProductNumber"))),
        ("Product Status", name(ci(product_status, "Status", "Name", "Value"))),
        ("Last Buy Date", str(ci(p, "DateLastBuyChance") or "")),
        ("Datasheet URL", normalise_url(ci(p, "DatasheetUrl", "DatasheetURL"))),
        ("Product URL", normalise_url(ci(p, "ProductUrl", "ProductURL"))),
        ("Product Image URL", normalise_url(ci(p, "PhotoUrl", "PhotoURL"))),
        ("Primary Video URL", normalise_url(ci(p, "PrimaryVideoUrl", "PrimaryVideoURL"))),
        ("RoHS Status", name(ci(classifications, "RohsStatus", "RoHSStatus"))),
        ("REACH Status", name(ci(classifications, "ReachStatus", "REACHStatus"))),
        ("Moisture Sensitivity Level", name(ci(classifications, "MoistureSensitivityLevel"))),
        ("ECCN", name(ci(classifications, "ExportControlClassNumber"))),
        ("HTSUS Code", name(ci(classifications, "HtsusCode", "HTSUSCode"))),
        ("Mounting Type", parameter_value(pa, "mounting type")),
        ("Package / Case", parameter_value(pa, "package / case", "package/case")),
        ("Supplier Device Package", parameter_value(pa, "supplier device package")),
        ("Size / Dimension", parameter_value(pa, "size / dimension", "size/dimension")),
        ("Height - Seated (Max)", parameter_value(pa, "height - seated (max)", "height (max)")),
        ("Operating Temperature", parameter_value(pa, "operating temperature")),
        ("Pin / Position Count", parameter_value(pa, "number of positions", "number of pins", "pin count")),
        ("Tolerance", parameter_value(pa, "tolerance", "frequency tolerance")),
        ("Voltage Rating", parameter_value(pa, "voltage - rated", "voltage rating", "voltage - dc reverse (vr) (max)", "drain to source voltage (vdss)")),
        ("Current Rating", parameter_value(pa, "current rating (amps)", "current rating", "current - output", "current - continuous drain (id) @ 25°c")),
        ("Power Rating", parameter_value(pa, "power (watts)", "power - max", "power dissipation (max)")),
        ("Provider", commercial_profile.get("provider", "")),
        ("Offer Count", len(commercial_offers(commercial_profile))),
        ("Provider Part Number", offer_value_summary(commercial_profile, "provider_part_number")),
        ("Currency", commercial_profile.get("provider_currency", "")),
        ("Pack Format", offer_value_summary(commercial_profile, "pack_format")),
        ("Packaging Code", offer_value_summary(commercial_profile, "packaging_code", "package_type")),
        ("Minimum Order Quantity", offer_value_summary(commercial_profile, "minimum_order_quantity")),
        ("Pack Quantity", offer_value_summary(commercial_profile, "pack_quantity")),
        ("Quantity Available", offer_value_summary(commercial_profile, "quantity_available")),
        ("Manufacturer Lead Weeks", commercial_profile.get("manufacturer_lead_weeks", "")),
        ("Additional Charge", all_additional_charges_summary(commercial_profile, "amount")),
        ("Additional Charge Description", all_additional_charges_summary(commercial_profile, "description")),
        ("Price Breaks", all_price_breaks_summary(commercial_profile)),
        ("Captured At UTC", record.captured_at_utc),
        ("Data Source Mode", record.source_mode),
        ("Data Provider", str(record.metadata.get("provider", "DigiKey"))),
    ])
    return values



def add_mapping_sheet(workbook, columns, sample_values):
    ws = workbook.create_sheet("Attribute Mapping")
    ws.append(["Group", "Workbook Column", "JSON Path / Source", "Sample Value", "Applicability", "Notes"])
    universal = {
        "Manufacturer", "Manufacturer Part Number", "Description", "Detailed Description",
        "Product Category", "Product Family", "Product Status", "Datasheet URL",
        "Product URL", "Product Image URL", "RoHS Status", "REACH Status", "ECCN",
        "HTSUS Code", "Captured At UTC", "Data Source Mode", "Data Provider",
    }
    for column in columns:
        applicability = "Universal" if column.key in universal else "Where available / commodity-specific"
        ws.append([
            column.group, column.heading, column.source, sample_values.get(column.key, ""),
            applicability, column.notes,
        ])
    format_reference_sheet(ws)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 52
    ws.column_dimensions["D"].width = 60
    ws.column_dimensions["E"].width = 34
    ws.column_dimensions["F"].width = 60



def classify_failure_status(error: Exception) -> str:
    """Map collection failures onto the four user-facing review states."""
    text = str(error).lower()
    if any(term in text for term in ("ambiguous", "multiple match", "multiple candidate")):
        return "Multiple Matches"
    if any(term in text for term in ("404", "not found", "no product", "no match")):
        return "Not Found"
    return "Review Required"

def _profile_attribute(profile, *aliases):
    attributes = (profile or {}).get("attributes") or {}
    lookup = {clean(key): value for key, value in attributes.items()}
    for alias in aliases:
        value = lookup.get(clean(alias))
        if value not in (None, "", "-"):
            return str(value)
    return ""


def _profile_compliance(profile, *aliases):
    compliance = (profile or {}).get("compliance") or {}
    lookup = {clean(key): value for key, value in compliance.items()}
    for alias in aliases:
        value = lookup.get(clean(alias))
        if value not in (None, "", "-"):
            return str(value)
    return ""


def _merged_custom_value(evidence, getter, attribute_name=""):
    """Collapse equivalent provider attributes; expose only genuine differences."""
    values = []
    for item in evidence:
        if not item.identity_match:
            continue
        value = str(getter(item.part_profile or {}) or "").strip()
        if value:
            key, display = normalise_attribute(value, attribute_name)
            values.append((item.provider, value, key, display))
    if not values:
        return ""
    keys = {key for _, _, key, _ in values}
    if len(keys) == 1:
        return values[0][3]
    return "EXCEPTION — " + "; ".join(f"{provider}: {raw}" for provider, raw, _, _ in values)


def _provider_evidence_text(evidence):
    """Legacy detailed provider summary retained for diagnostics/tests.

    Operational workbook output uses _provider_results_text().
    """
    return "\n".join(f"{item.provider}: {_concise_provider_status(item)}" for item in evidence)


def _provider_results_text(evidence):
    """Compact provider evidence counts suitable for large provider sets."""
    counts = {"matched": 0, "not listed": 0, "error": 0, "unconfirmed": 0, "skipped": 0}
    for item in evidence:
        if item.identity_match:
            counts["matched"] += 1
            continue
        status = _normalised_provider_status(item.execution_status, item.message)
        if status == "no_match":
            counts["not listed"] += 1
        elif status == "error":
            counts["error"] += 1
        elif status == "skipped":
            counts["skipped"] += 1
        elif status == "success":
            counts["unconfirmed"] += 1
        else:
            counts["error"] += 1
    order = ("matched", "not listed", "error", "unconfirmed", "skipped")
    return "; ".join(f"{counts[label]} {label}" for label in order if counts[label]) or "No provider result"


def _provider_price_summary(profile):
    return all_price_breaks_summary(profile) if profile else ""


def _provider_profile(evidence, provider_name):
    for item in evidence:
        if item.provider.casefold() == provider_name.casefold():
            return item.commercial_profile or {}
    return {}


def _commercial_profile_has_useful_data(profile):
    """True when a provider has something worth occupying a dashboard block."""
    profile = profile or {}
    if provider_availability(profile) not in (None, ""):
        return True
    if lead_time_display(profile.get("manufacturer_lead_weeks", "")):
        return True
    if str(profile.get("provider_currency") or "").strip():
        return True
    if _provider_price_summary(profile):
        return True
    return False


def _provider_dashboard_profiles(evidence, limit=3):
    """Return provider+commercial-profile pairs, skipping empty provider blocks.

    Registration/evidence order is preserved for now.  This is intentionally
    not a commercial ranking policy; later releases can rank the providers that
    actually returned useful data.
    """
    rows = []
    for item in evidence:
        profile = item.commercial_profile or {}
        if _commercial_profile_has_useful_data(profile):
            rows.append((item.provider, profile))
        if len(rows) >= limit:
            break
    return rows


def _apply_provider_dashboard(values, evidence):
    for position in range(1, 4):
        values[f"Provider #{position} Name"] = ""
        values[f"Provider #{position} Available"] = ""
        values[f"Provider #{position} Lead Time"] = ""
        values[f"Provider #{position} Currency"] = ""
        values[f"Provider #{position} Price Breaks"] = ""
    for position, (provider, profile) in enumerate(_provider_dashboard_profiles(evidence), 1):
        values[f"Provider #{position} Name"] = provider
        values[f"Provider #{position} Available"] = provider_availability(profile)
        lead = normalise_lead_time(profile.get("manufacturer_lead_weeks", ""))
        values[f"Provider #{position} Lead Time"] = lead.weeks if lead.weeks is not None else lead.display
        values[f"Provider #{position} Currency"] = profile.get("provider_currency", "")
        values[f"Provider #{position} Price Breaks"] = _provider_price_summary(profile)


def build_combined_result(row, requested_mfg, requested_mpn, evidence, *, bom_context=None, local_context=None, recovered_mpn="", candidate_count=0):
    bom_context = bom_context or {}
    local_context = local_context or PartsMasterLookup().find(requested_mfg, requested_mpn)
    status, reason = evidence_status(evidence)
    matched = [item.provider for item in evidence if item.identity_match]
    values = OrderedDict((column.key, "") for column in enriched_columns())
    lifecycle = merged_value(evidence, "lifecycle_status")
    provider_results = _provider_results_text(evidence)
    providers_matched = ", ".join(matched)
    values.update({
        "Source Row": row,
        "Requested Manufacturer": requested_mfg,
        "Requested MPN": requested_mpn,
        "Match Status": status,
        "Reason": reason,
        "Provider Results": provider_results,
        "Engineering Confirmation": engineering_confirmation(evidence),
        "Review Observation": provider_review_observation(
            match_status=status, provider_results=provider_results,
            providers_matched=providers_matched, local_context=local_context, lifecycle=lifecycle,
            requested_mpn=requested_mpn, recovered_mpn=recovered_mpn, candidate_count=candidate_count,
        ),
        "BOM Description": bom_context.get("description", ""),
        "BOM Value": bom_context.get("value", ""),
        "BOM Footprint": bom_context.get("footprint", ""),
        "BOM Quantity": bom_context.get("quantity", ""),
        "BOM DNP": bom_context.get("dnp", ""),
        "Local Knowledge Status": local_context.status,
        "AIPN": local_context.aipn,
        "Local Lifecycle": local_context.lifecycle,
        "Datasheet Evidence Status": local_context.datasheet_status,
        "Static Datasheet": local_context.datasheet_local_file,
        "Manufacturer": merged_value(evidence, "manufacturer") or local_context.manufacturer,
        "Manufacturer Part Number": merged_value(evidence, "manufacturer_part_number") or local_context.mpn,
        "Description": merged_value(evidence, "description") or local_context.description,
        "Detailed Description": merged_value(evidence, "detailed_description"),
        "Product Status": lifecycle or local_context.lifecycle,
        "Mounting Type": merged_value(evidence, "mounting_type"),
        "Package / Case": merged_value(evidence, "package"),
        "Operating Temperature": _merged_custom_value(evidence, lambda p: _profile_attribute(p, "Operating Temperature"), "Operating Temperature"),
        "Pin / Position Count": _merged_custom_value(evidence, lambda p: _profile_attribute(p, "Number of Positions", "Number of Pins", "Pin Count"), "Pin / Position Count"),
        "Tolerance": _merged_custom_value(evidence, lambda p: _profile_attribute(p, "Tolerance", "Frequency Tolerance"), "Tolerance"),
        "Voltage Rating": _merged_custom_value(evidence, lambda p: _profile_attribute(p, "Voltage - Rated", "Voltage Rating", "Voltage - DC Reverse (Vr) (Max)"), "Voltage Rating"),
        "Current Rating": _merged_custom_value(evidence, lambda p: _profile_attribute(p, "Current Rating", "Current - Output", "Current - Continuous Drain"), "Current Rating"),
        "Power Rating": _merged_custom_value(evidence, lambda p: _profile_attribute(p, "Power (Watts)", "Power - Max", "Power Dissipation"), "Power Rating"),
        "Datasheet URL": merged_value(evidence, "datasheet_url") or local_context.datasheet_active_url,
        "Product URL": merged_value(evidence, "product_url"),
        "Product Image URL": merged_value(evidence, "image_url"),
        "RoHS Status": merged_value(evidence, "rohs_status"),
        "REACH Status": _merged_custom_value(evidence, lambda p: _profile_compliance(p, "REACH", "REACH Status")),
        "ECCN": _merged_custom_value(evidence, lambda p: _profile_compliance(p, "ECCN", "Export Control Class Number")),
        "HTSUS Code": _merged_custom_value(evidence, lambda p: _profile_compliance(p, "HTSUS", "HTSUS Code")),
    })
    _apply_provider_dashboard(values, evidence)
    return values


def run(args):
    run_started = perf_counter()
    _progress(f"PDC v{APP_VERSION}: opening {args.input}")
    source = load_input_workbook(args.input, data_only=False)
    input_sheet = source[args.sheet] if args.sheet else source.active
    headers = [cell.value for cell in input_sheet[1]]
    manufacturer_column, mpn_column = locate_columns(headers)
    description_column = locate_optional_column(headers, DESCRIPTION)
    value_column = locate_optional_column(headers, VALUE)
    footprint_column = locate_optional_column(headers, FOOTPRINT)
    quantity_column = locate_optional_column(headers, QUANTITY)
    dnp_column = locate_optional_column(headers, DNP)

    input_rows = []
    for row_number in range(max(2, args.start_row), input_sheet.max_row + 1):
        manufacturer = str(input_sheet.cell(row_number, manufacturer_column).value or "").strip()
        mpn = str(input_sheet.cell(row_number, mpn_column).value or "").strip()
        if manufacturer or mpn:
            input_rows.append((
                row_number, manufacturer, mpn, {
                    "description": input_cell(input_sheet, row_number, description_column),
                    "value": input_cell(input_sheet, row_number, value_column),
                    "footprint": input_cell(input_sheet, row_number, footprint_column),
                    "quantity": input_cell(input_sheet, row_number, quantity_column),
                    "dnp": input_cell(input_sheet, row_number, dnp_column),
                }
            ))
        if args.max_parts and len(input_rows) >= args.max_parts:
            break

    settings = Settings.from_env()
    knowledge_base = KnowledgeBaseManager()
    parts_master_lookup = PartsMasterLookup()
    digikey = DigiKeyProvider(settings, knowledge_base)
    mouser = MouserProvider(MouserSettings.from_env(), knowledge_base)
    tme = TmeProvider(TmeSettings.from_env(), knowledge_base)
    provider_manager = ProviderManager([digikey, mouser, tme])
    _progress(
        f"PDC v{APP_VERSION}: loaded {len(input_rows)} parts; "
        f"providers={', '.join(provider_manager.names)}; "
        f"DigiKey site={settings.site}, currency={settings.currency}"
    )
    if args.validate_only:
        return

    results = []
    commercial_rows = []
    identity_candidate_rows = []
    _progress("Initialising provider data...")
    manufacturer_result = provider_manager.execute(digikey, "manufacturers", args.force_refresh)
    _progress("Provider initialisation complete.")
    manufacturer_catalogue = manufacturer_result.data if manufacturer_result.succeeded else None

    for index, (row_number, manufacturer, mpn, bom_context) in enumerate(input_rows, 1):
        _progress(
            f"[{index}/{len(input_rows)}] "
            f"{manufacturer or '<MFG missing>'} {mpn or '<MPN missing>'}"
        )
        evidence = []
        resolved = None
        candidate_evidence = []
        recovered = recover_mpn_from_bom(manufacturer, mpn, bom_context)
        search_mpn = mpn or (recovered.mpn if recovered else "")
        if recovered:
            candidate_evidence.append(recovered)
            _progress(f"    Identity recovery: BOM context candidate {recovered.mpn}")

        if not search_mpn:
            for provider_name in provider_manager.names:
                evidence.append(
                    ProviderEvidence(
                        provider_name,
                        "skipped",
                        "MPN missing and no BOM-context candidate found; provider query not attempted.",
                    )
                )
                _progress(f"    {provider_name}: skipped - MPN missing")
        else:
            # Resolve DigiKey manufacturer locally, then collect independent provider
            # detail calls concurrently.  Provider result interpretation remains
            # deterministic and provider-neutral after collection completes.
            provider_result = None
            digikey_resolution_error = None
            if manufacturer_catalogue is not None:
                try:
                    resolved = resolve_manufacturer(manufacturer, manufacturer_catalogue)
                    if resolved.manufacturer_id is None:
                        raise RuntimeError(
                            f"Manufacturer resolution {resolved.status}: {resolved.reason} "
                            f"(best={resolved.matched_name!r}, confidence={resolved.confidence:.2f})"
                        )
                except Exception as error:
                    digikey_resolution_error = error

            futures = {}
            with ThreadPoolExecutor(max_workers=3, thread_name_prefix="pdc-provider") as pool:
                if manufacturer_catalogue is not None and digikey_resolution_error is None:
                    futures["DigiKey"] = pool.submit(
                        provider_manager.execute, digikey, "details", search_mpn, resolved.manufacturer_id, args.force_refresh,
                        input_manufacturer=manufacturer, resolved_manufacturer=resolved.matched_name,
                    )
                futures["Mouser"] = pool.submit(
                    provider_manager.execute, mouser, "details", search_mpn, None, args.force_refresh,
                    input_manufacturer=manufacturer,
                    resolved_manufacturer=(resolved.matched_name if resolved else manufacturer),
                )
                futures["TME"] = pool.submit(
                    provider_manager.execute, tme, "details", search_mpn, None, args.force_refresh,
                    input_manufacturer=manufacturer,
                    resolved_manufacturer=(resolved.matched_name if resolved else manufacturer),
                )
                concurrent_results = {name: future.result() for name, future in futures.items()}

            if manufacturer_catalogue is None:
                provider_result = manufacturer_result
            elif digikey_resolution_error is not None:
                evidence.append(ProviderEvidence("DigiKey", "error", str(digikey_resolution_error)))
            else:
                provider_result = concurrent_results.get("DigiKey")

            if provider_result is not None:
                if provider_result.succeeded:
                    record = provider_result.data
                    profile = dict(record.part_profile or {})
                    profile["identity_match"] = provider_identity_match(manufacturer, mpn, profile) if mpn else False
                    if recovered and normalise_mpn(profile.get("manufacturer_part_number")) == normalise_mpn(search_mpn):
                        candidate_evidence.append(RecoveryCandidate(
                            manufacturer=profile.get("manufacturer", manufacturer),
                            mpn=profile.get("manufacturer_part_number", search_mpn),
                            relationship="Recovered MPN candidate", sources=("DigiKey",),
                            package=profile.get("package", ""), datasheet_url=profile.get("datasheet_url", ""),
                            footprint_check=footprint_consistency(
                                bom_context.get("footprint", ""), profile.get("package", ""),
                                profile.get("manufacturer_part_number", ""),
                            ),
                            notes="Provider supports BOM-context MPN candidate; review required.",
                        ))
                    elif mpn:
                        candidate = candidate_from_profile(
                            provider="DigiKey", requested_manufacturer=manufacturer, reference_mpn=mpn,
                            profile=profile, bom_footprint=bom_context.get("footprint", ""),
                        )
                        if candidate:
                            candidate_evidence.append(candidate)
                    candidate_evidence.extend(discover_payload_candidates(
                        "DigiKey", record.provider_response, requested_manufacturer=manufacturer,
                        reference_mpn=search_mpn, bom_footprint=bom_context.get("footprint", ""),
                    ))
                    execution_status = "success" if profile.get("manufacturer_part_number") else "no_match"
                    evidence.append(ProviderEvidence(
                        "DigiKey", execution_status, part_profile=profile,
                        commercial_profile=record.commercial_profile,
                        captured_at_utc=record.captured_at_utc, source_mode=record.source_mode,
                    ))
                    commercial_rows.extend(commercial_analysis_rows(
                        {"Source Row": row_number, "Manufacturer": profile.get("manufacturer", ""),
                         "Manufacturer Part Number": profile.get("manufacturer_part_number", "")},
                        record.commercial_profile,
                    ))
                else:
                    normal_status = _normalised_provider_status(
                        provider_result.status.value, provider_result.message or "",
                    )
                    evidence.append(ProviderEvidence("DigiKey", normal_status, provider_result.message or ""))
            digikey_item = next((item for item in evidence if item.provider == "DigiKey"), None)
            if digikey_item:
                _progress(f"    DigiKey: {_concise_provider_status(digikey_item)}")

            # Mouser
            mouser_result = concurrent_results["Mouser"]
            if mouser_result.succeeded:
                record = mouser_result.data
                profile = dict(record.part_profile or {})
                profile["identity_match"] = provider_identity_match(manufacturer, mpn, profile) if mpn else False
                if recovered and normalise_mpn(profile.get("manufacturer_part_number")) == normalise_mpn(search_mpn):
                    candidate_evidence.append(RecoveryCandidate(
                        manufacturer=profile.get("manufacturer", manufacturer),
                        mpn=profile.get("manufacturer_part_number", search_mpn),
                        relationship="Recovered MPN candidate", sources=("Mouser",),
                        package=profile.get("package", ""), datasheet_url=profile.get("datasheet_url", ""),
                        footprint_check=footprint_consistency(
                            bom_context.get("footprint", ""), profile.get("package", ""),
                            profile.get("manufacturer_part_number", ""),
                        ),
                        notes="Provider supports BOM-context MPN candidate; review required.",
                    ))
                elif mpn:
                    candidate = candidate_from_profile(
                        provider="Mouser", requested_manufacturer=manufacturer, reference_mpn=mpn,
                        profile=profile, bom_footprint=bom_context.get("footprint", ""),
                    )
                    if candidate:
                        candidate_evidence.append(candidate)
                candidate_evidence.extend(discover_payload_candidates(
                    "Mouser", record.provider_response, requested_manufacturer=manufacturer,
                    reference_mpn=search_mpn, bom_footprint=bom_context.get("footprint", ""),
                ))
                execution_status = "success" if profile.get("manufacturer_part_number") else "no_match"
                evidence.append(ProviderEvidence(
                    "Mouser", execution_status, part_profile=profile,
                    commercial_profile=record.commercial_profile,
                    captured_at_utc=record.captured_at_utc, source_mode=record.source_mode,
                ))
                commercial_rows.extend(commercial_analysis_rows(
                    {"Source Row": row_number, "Manufacturer": profile.get("manufacturer", ""),
                     "Manufacturer Part Number": profile.get("manufacturer_part_number", "")},
                    record.commercial_profile,
                ))
            else:
                normal_status = _normalised_provider_status(mouser_result.status.value, mouser_result.message or "")
                evidence.append(ProviderEvidence("Mouser", normal_status, mouser_result.message or ""))
            mouser_item = next((item for item in evidence if item.provider == "Mouser"), None)
            if mouser_item:
                _progress(f"    Mouser: {_concise_provider_status(mouser_item)}")

            # TME
            tme_result = concurrent_results["TME"]
            if tme_result.succeeded:
                record = tme_result.data
                profile = dict(record.part_profile or {})
                profile["identity_match"] = provider_identity_match(manufacturer, mpn, profile) if mpn else False
                if recovered and normalise_mpn(profile.get("manufacturer_part_number")) == normalise_mpn(search_mpn):
                    candidate_evidence.append(RecoveryCandidate(
                        manufacturer=profile.get("manufacturer", manufacturer),
                        mpn=profile.get("manufacturer_part_number", search_mpn),
                        relationship="Recovered MPN candidate", sources=("TME",),
                        package=profile.get("package", ""), datasheet_url=profile.get("datasheet_url", ""),
                        footprint_check=footprint_consistency(
                            bom_context.get("footprint", ""), profile.get("package", ""),
                            profile.get("manufacturer_part_number", ""),
                        ),
                        notes="Provider supports BOM-context MPN candidate; review required.",
                    ))
                elif mpn:
                    candidate = candidate_from_profile(
                        provider="TME", requested_manufacturer=manufacturer, reference_mpn=mpn,
                        profile=profile, bom_footprint=bom_context.get("footprint", ""),
                    )
                    if candidate:
                        candidate_evidence.append(candidate)
                candidate_evidence.extend(discover_payload_candidates(
                    "TME", record.provider_response, requested_manufacturer=manufacturer,
                    reference_mpn=search_mpn, bom_footprint=bom_context.get("footprint", ""),
                ))
                execution_status = "success" if profile.get("manufacturer_part_number") else "no_match"
                evidence.append(ProviderEvidence(
                    "TME", execution_status, part_profile=profile,
                    commercial_profile=record.commercial_profile,
                    captured_at_utc=record.captured_at_utc, source_mode=record.source_mode,
                ))
                commercial_rows.extend(commercial_analysis_rows(
                    {"Source Row": row_number, "Manufacturer": profile.get("manufacturer", ""),
                     "Manufacturer Part Number": profile.get("manufacturer_part_number", "")},
                    record.commercial_profile,
                ))
            else:
                normal_status = _normalised_provider_status(tme_result.status.value, tme_result.message or "")
                evidence.append(ProviderEvidence("TME", normal_status, tme_result.message or ""))
            tme_item = next((item for item in evidence if item.provider == "TME"), None)
            if tme_item:
                _progress(f"    TME: {_concise_provider_status(tme_item)}")

        # 4.7.2c discovery state machine.
        # Exhaust exact-format and alphanumeric discovery before family broadening.
        # A same-manufacturer candidate whose alphanumeric MPN equals the source
        # search key is a strong formatting-normalised identity candidate.
        if search_mpn:
            def _strong_format_candidate(items):
                wanted = normalise_mpn(search_mpn)
                return next((
                    c for c in consolidate_candidates(items)
                    if c.mpn and normalise_mpn(c.mpn) == wanted
                    and c.relationship != "Manufacturer family discovery candidate"
                ), None)

            # DigiKey keyword discovery must try BOTH the source text and the
            # punctuation-free key; Product Details alone is intentionally strict.
            for variant in search_variants(search_mpn):
                _progress(f"    Discovery search: {variant.kind} = {variant.text}")
                dk = provider_manager.execute(digikey, "keyword_search", variant.text, record_count=25)
                if dk.succeeded:
                    payload = dk.data[0] if isinstance(dk.data, tuple) else dk.data
                    candidate_evidence.extend(discover_payload_candidates(
                        "DigiKey", payload, requested_manufacturer=manufacturer,
                        reference_mpn=search_mpn, bom_footprint=bom_context.get("footprint", ""),
                    ))

                # Do not duplicate the source-text detail call already performed
                # during the first provider pass.  The alphanumeric representation
                # is a genuine second discovery form for Mouser/TME.
                if variant.kind != "Source Search Text":
                    for provider_name, provider in (("Mouser", mouser), ("TME", tme)):
                        retry = provider_manager.execute(
                            provider, "details", variant.text, None, args.force_refresh,
                            input_manufacturer=manufacturer,
                            resolved_manufacturer=(resolved.matched_name if resolved else manufacturer),
                        )
                        if retry.succeeded:
                            record = retry.data
                            profile = dict(record.part_profile or {})
                            candidate = candidate_from_profile(
                                provider=provider_name, requested_manufacturer=manufacturer,
                                reference_mpn=search_mpn, profile=profile,
                                bom_footprint=bom_context.get("footprint", ""),
                            )
                            if candidate:
                                candidate_evidence.append(candidate)
                            candidate_evidence.extend(discover_payload_candidates(
                                provider_name, record.provider_response,
                                requested_manufacturer=manufacturer, reference_mpn=search_mpn,
                                bom_footprint=bom_context.get("footprint", ""),
                            ))

                strong = _strong_format_candidate(candidate_evidence)
                if strong:
                    _progress(
                        f"    Strong formatting-normalised candidate: {strong.mpn} "
                        f"({', '.join(strong.sources)}) - family broadening stopped"
                    )
                    break

            strong = _strong_format_candidate(candidate_evidence)

            # Only family-search when exact/alphanumeric discovery produced no
            # strong formatting-equivalent candidate.
            if not strong:
                for family_variant in family_search_variants(search_mpn):
                    _progress(
                        f"    Family discovery: {family_variant.text} "
                        f"({family_variant.reduction_percent}% right-side reduction)"
                    )
                    level_candidates = []
                    dk = provider_manager.execute(digikey, "keyword_search", family_variant.text, record_count=25)
                    if dk.succeeded:
                        payload = dk.data[0] if isinstance(dk.data, tuple) else dk.data
                        level_candidates.extend(discover_payload_candidates(
                            "DigiKey", payload, requested_manufacturer=manufacturer,
                            reference_mpn=search_mpn, bom_footprint=bom_context.get("footprint", ""),
                        ))
                    for provider_name, provider in (("Mouser", mouser), ("TME", tme)):
                        retry = provider_manager.execute(
                            provider, "details", family_variant.text, None, args.force_refresh,
                            input_manufacturer=manufacturer,
                            resolved_manufacturer=(resolved.matched_name if resolved else manufacturer),
                        )
                        if retry.succeeded:
                            record = retry.data
                            level_candidates.extend(discover_payload_candidates(
                                provider_name, record.provider_response,
                                requested_manufacturer=manufacturer, reference_mpn=search_mpn,
                                bom_footprint=bom_context.get("footprint", ""),
                            ))
                    if level_candidates:
                        candidate_evidence.extend([
                            RecoveryCandidate(
                                manufacturer=c.manufacturer, mpn=c.mpn,
                                relationship="Manufacturer family discovery candidate",
                                sources=c.sources, package=c.package, datasheet_url=c.datasheet_url,
                                footprint_check=c.footprint_check,
                                notes=(
                                    f"Discovered from family key {family_variant.text} after "
                                    f"{family_variant.reduction_percent}% right-side reduction. "
                                    "Candidate requires manufacturer-document/order-table verification."
                                ),
                            ) for c in level_candidates
                        ])
                        break

            # Cross-provider learning uses the strongest discovered proper MPN,
            # including formatting-normalised identity candidates.
            provisional = consolidate_candidates(candidate_evidence)
            learned = _strong_format_candidate(candidate_evidence) or next((
                c for c in provisional
                if c.mpn and c.relationship not in (
                    "Manufacturer family discovery candidate", "Recovered MPN candidate"
                )
            ), None)
            if learned:
                _progress(f"    Cross-provider retry using discovered MPN: {learned.mpn}")
                confirmed = {item.provider for item in evidence if item.identity_match}
                for provider_name, provider in (("DigiKey", digikey), ("Mouser", mouser), ("TME", tme)):
                    if provider_name in confirmed:
                        continue
                    if provider_name == "DigiKey":
                        if resolved is None or resolved.manufacturer_id is None:
                            continue
                        retry = provider_manager.execute(
                            digikey, "details", learned.mpn, resolved.manufacturer_id, args.force_refresh,
                            input_manufacturer=manufacturer, resolved_manufacturer=resolved.matched_name,
                        )
                    else:
                        retry = provider_manager.execute(
                            provider, "details", learned.mpn, None, args.force_refresh,
                            input_manufacturer=manufacturer,
                            resolved_manufacturer=(resolved.matched_name if resolved else manufacturer),
                        )
                    if retry.succeeded:
                        record = retry.data
                        profile = dict(record.part_profile or {})
                        if (
                            names_equivalent(manufacturer, profile.get("manufacturer", ""))
                            and normalise_mpn(profile.get("manufacturer_part_number"))
                            == normalise_mpn(learned.mpn)
                        ):
                            candidate_evidence.append(RecoveryCandidate(
                                manufacturer=profile.get("manufacturer", manufacturer),
                                mpn=profile.get("manufacturer_part_number", learned.mpn),
                                relationship="Cross-provider confirmed candidate",
                                sources=(provider_name,), package=profile.get("package", ""),
                                datasheet_url=profile.get("datasheet_url", ""),
                                footprint_check=footprint_consistency(
                                    bom_context.get("footprint", ""), profile.get("package", ""),
                                    profile.get("manufacturer_part_number", ""),
                                ),
                                notes="Provider confirmed an MPN first resolved through discovery; review required.",
                            ))

        candidates = consolidate_candidates(candidate_evidence)
        for candidate in candidates:
            identity_candidate_rows.append(OrderedDict([
                ("Source Row", row_number),
                ("Input Manufacturer", manufacturer),
                ("Input MPN", mpn),
                ("BOM Value", bom_context.get("value", "")),
                ("BOM Footprint", bom_context.get("footprint", "")),
                ("Candidate Manufacturer", candidate.manufacturer),
                ("Candidate MPN", candidate.mpn),
                ("Relationship", candidate.relationship),
                ("Evidence Sources", ", ".join(candidate.sources)),
                ("Candidate Package / Case", candidate.package),
                ("Footprint Check", candidate.footprint_check),
                ("Datasheet URL", candidate.datasheet_url),
                ("Status", "Review Required"),
                ("Notes", candidate.notes),
            ]))

        local_context = parts_master_lookup.find(manufacturer, mpn)
        result = build_combined_result(
            row_number, manufacturer, mpn, evidence,
            bom_context=bom_context, local_context=local_context,
            recovered_mpn=(recovered.mpn if recovered else ""), candidate_count=len(candidates),
        )
        results.append(result)
        _progress(f"    Review: {result.get('Match Status', '')} - {result.get('Review Observation', '')}")

        if search_mpn:
            knowledge_base.save_part_summary(
                manufacturer=manufacturer,
                mpn=search_mpn,
                summary={
                    "input_mpn": mpn,
                    "identity_recovery_source": (", ".join(recovered.sources) if recovered else ""),
                    "match_status": result.get("Match Status", ""),
                    "reason": result.get("Reason", ""),
                    "providers_matched": [item.provider for item in evidence if item.identity_match],
                    "engineering_confirmation": result.get("Engineering Confirmation", ""),
                    "identity_candidates": [
                        {
                            "manufacturer": candidate.manufacturer, "mpn": candidate.mpn,
                            "relationship": candidate.relationship, "sources": list(candidate.sources),
                            "package": candidate.package, "footprint_check": candidate.footprint_check,
                            "datasheet_url": candidate.datasheet_url,
                        } for candidate in candidates
                    ],
                    "providers": [
                        {
                            "provider": item.provider,
                            "execution_status": item.execution_status,
                            "identity_match": item.identity_match,
                            "message": item.message,
                            "captured_at_utc": item.captured_at_utc,
                            "source_mode": item.source_mode,
                        }
                        for item in evidence
                    ],
                },
            )

    columns = enriched_columns()
    keys = column_keys(columns)
    headings = display_headings(columns)
    reason_column = WorkbookColumn(
        "Status", "Reason", "Reason",
        "PDC combined provider evidence", "Detailed provider evidence"
    )
    review_columns = columns[:4] + [reason_column] + columns[4:]
    review_keys = column_keys(review_columns)
    review_headings = display_headings(review_columns)

    output = Workbook()
    enriched = output.active
    enriched.title = "Enriched Parts"
    add_group_headers(enriched, columns)
    enriched.append(headings)
    for result in results:
        enriched.append([result.get(key, "") for key in keys])

    # 4.7.2c: Review Required is an exception queue, not a duplicate Enriched sheet.
    review = output.create_sheet("Review Required")
    review_headings = [
        "Source Row", "Manufacturer", "Source MPN / Search Field", "Review Reason",
        "Recovered / Suggested MPN", "Identity Status", "Provider Results",
        "Attribute Exceptions", "BOM Footprint", "Footprint Check",
        "Datasheet Evidence", "Engineering Action / Decision",
    ]
    review.append(review_headings)
    candidates_by_row = {}
    for candidate_row in identity_candidate_rows:
        candidates_by_row.setdefault(candidate_row.get("Source Row"), []).append(candidate_row)

    for result in results:
        observation = str(result.get("Review Observation", "") or "")
        attribute_exceptions = "; ".join(
            f"{key}: {value}" for key, value in result.items()
            if isinstance(value, str) and value.startswith("EXCEPTION")
        )
        needs_review = (
            result.get("Match Status") != "Matched"
            or bool(attribute_exceptions)
            or "review" in observation.casefold()
            or "exception" in observation.casefold()
        )
        if not needs_review:
            continue
        row_candidates = candidates_by_row.get(result.get("Source Row"), [])
        suggested = "; ".join(dict.fromkeys(
            str(c.get("Candidate MPN", "") or "") for c in row_candidates
            if c.get("Candidate MPN")
        ))
        footprint_checks = "; ".join(dict.fromkeys(
            str(c.get("Footprint Check", "") or "") for c in row_candidates
            if c.get("Footprint Check")
        ))
        source_search = result.get("Requested MPN", "") or result.get("BOM Value", "")
        review.append([
            result.get("Source Row", ""),
            result.get("Manufacturer", "") or result.get("Requested Manufacturer", ""),
            source_search,
            observation or result.get("Reason", ""),
            suggested,
            result.get("Match Status", ""),
            result.get("Provider Results", ""),
            attribute_exceptions,
            result.get("BOM Footprint", ""),
            footprint_checks,
            result.get("Datasheet Evidence", "") or result.get("Datasheet Evidence Status", ""),
            "",
        ])

    commercial = output.create_sheet("Commercial Analysis")
    commercial_headings = [
        "Source Row", "Manufacturer", "Manufacturer Part Number", "Provider",
        "Provider Part Number", "Currency", "Pack Format", "Packaging Code",
        "Minimum Order Quantity", "Pack Quantity", "Quantity Available",
        "Manufacturer Lead Weeks", "Break Quantity", "Unit Price",
        "Extended Price", "Additional Charge", "Additional Charge Currency",
        "Additional Charge Description", "Additional Charge Application",
        "Captured At UTC",
    ]
    commercial.append(commercial_headings)
    commercial_rows.sort(key=lambda row: (
        str(row.get("Manufacturer", "")).upper(),
        str(row.get("Manufacturer Part Number", "")).upper(),
        str(row.get("Provider", "")).upper(),
        str(row.get("Pack Format", "")).upper(),
        _numeric_sort(row.get("Break Quantity")),
    ))
    for row in commercial_rows:
        commercial.append([row.get(heading, "") for heading in commercial_headings])

    candidates_sheet = output.create_sheet("Identity Candidates")
    candidate_headings = [
        "Source Row", "Input Manufacturer", "Input MPN", "BOM Value", "BOM Footprint",
        "Candidate Manufacturer", "Candidate MPN", "Relationship", "Evidence Sources",
        "Candidate Package / Case", "Footprint Check", "Datasheet URL", "Status", "Notes",
    ]
    candidates_sheet.append(candidate_headings)
    for candidate_row in identity_candidate_rows:
        candidates_sheet.append([candidate_row.get(heading, "") for heading in candidate_headings])
    format_reference_sheet(candidates_sheet)
    candidates_sheet.freeze_panes = "A2"

    summary = output.create_sheet("BOM Review Summary", 0)
    summary.append(["PDC Operational BOM Review", "Value"])
    for label, value in summary_rows(results):
        summary.append([label, value])
    summary.append([])
    summary.append(["Identity Candidates Discovered", len(identity_candidate_rows)])
    summary.append(["Providers", ", ".join(provider_manager.names)])
    for label, value in provider_manager.diagnostic_rows():
        summary.append([label, value])
    summary.append(["Review Rule", "DNP rows are reviewed like fitted rows; DNP remains assembly context."])
    summary.append(["Approval", "PDC reports evidence and exceptions only; no automatic engineering approval."])
    format_reference_sheet(summary)
    summary.column_dimensions["A"].width = 42
    summary.column_dimensions["B"].width = 90

    format_review_sheet(enriched, headings)
    format_reference_sheet(review)
    review.freeze_panes = "A2"
    review.auto_filter.ref = review.dimensions
    review.column_dimensions["D"].width = 48
    review.column_dimensions["E"].width = 34
    review.column_dimensions["G"].width = 28
    review.column_dimensions["H"].width = 48
    review.column_dimensions["L"].width = 36
    format_reference_sheet(commercial)

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)
    output.save(args.output)
    elapsed = perf_counter() - run_started
    _progress("")
    _progress("Run diagnostics:")
    for label, value in provider_manager.diagnostic_rows():
        _progress(f"    {label}: {value}")
    _progress(f"    Total elapsed: {elapsed:.1f} s")
    print("Created", args.output, flush=True)


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", default="output/AIPN_Enriched.xlsx")
    parser.add_argument("--sheet")
    parser.add_argument("--start-row", type=int, default=2)
    parser.add_argument("--max-parts", type=int)
    parser.add_argument("--force-refresh", action="store_true")
    parser.add_argument("--validate-only", action="store_true")
    run(parser.parse_args())
